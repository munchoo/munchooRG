#%%
import imp
from matplotlib import offsetbox
import openpyxl
from openpyxl.drawing.image import Image as im3
import glob
import os
import tkinter
from tkinter import Image, filedialog
import pandas as pd
from PIL import Image as pi

path = os.getcwd()
Ofc_list = pd.read_excel('C:/Users/Administrator/Desktop/GSpay홍보물취합_자동화/output.xlsx')

# 워크시트 불러오기 tkinter
# root = tkinter.Tk()
# root.withdraw()
# dir_path = filedialog.askopenfile(filetypes=[('엑셀','*.xlsx')])


def resizeImg(size, img_path):
    #이미지 객체 생성
    img = pi.open(img_path)
    #이미지 resize
    resize_img = img.resize(size)
    resize_img = resize_img.rotate(270, expand=True)
    #수정한 이미지 저장
    resize_img.save(img_path, "JPEG", quality=95)


dir_path = 'C:/Users/Administrator/Desktop/GSpay홍보물취합_자동화/source.xlsx'
save_path2 = 'C:/Users/Administrator/Desktop/GSpay홍보물취합_자동화/source2.xlsx'
img_path = 'C:/Users/Administrator/Desktop/GSpay홍보물취합_자동화/img/'
wb = openpyxl.load_workbook(dir_path)

img_height = 247.6
img_width = 312.8
size = (400,200)

row_num = 5
col_num = 3

source = wb[wb.sheetnames[0]]

# Uniq_team = Ofc_list['점포(팀)'].unique()
a = Ofc_list[Ofc_list['점포(팀)'] == '영업6팀']
unique_name = a['파트명'].unique()

# 파트별 시트만들기
for i in unique_name:
    target = wb.copy_worksheet(source)
    target.title = i
    c = a[a['파트명']==i]
    row_num = 5
# 시트에 사진넣기
    for idx, row in c.iterrows():
        col_num = 3
        try:
            for i in range(1,4):
                try:
                    resizeImg(size, f'{img_path}{row["점포명"]}_{i}.jpg')
                    img = im3(f'{img_path}{row["점포명"]}_{i}.jpg')

                    img.height = img_height
                    img.width = img_width


                    #점포명과 이미지 넣기
                    target.cell(row=row_num, column=col_num, value=row['점포명'])
                    colletter =target.cell(row=row_num, column=col_num+i).column_letter
                    target.add_image(img, f'{colletter}{row_num}')

                    print('success')
                except:
                    pass
             
        
            row_num += 1
        except:
            row_num += 1
            print('fail')
            continue

wb.remove_sheet(wb['원시'])
wb.save(save_path2)